"""
Data export utilities
"""

import json
import csv
import asyncio
import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)

class DataExporter:
    """Export scan data to various formats"""
    
    def __init__(self, database_manager):
        self.db = database_manager
    
    async def export_json(self, filename: str, filters: Optional[Dict[str, Any]] = None) -> bool:
        """Export data to JSON format"""
        try:
            data = await self._get_export_data(filters)
            
            export_data = {
                'export_info': {
                    'timestamp': datetime.utcnow().isoformat(),
                    'format': 'json',
                    'total_records': len(data),
                    'filters': filters or {}
                },
                'servers': data
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"Exported {len(data)} servers to JSON: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            return False
    
    async def export_csv(self, filename: str, filters: Optional[Dict[str, Any]] = None) -> bool:
        """Export data to CSV format"""
        try:
            data = await self._get_export_data(filters)
            
            if not data:
                logger.warning("No data to export")
                return False
            
            # Define CSV columns
            columns = [
                'ip', 'port', 'minecraft_version', 'server_software', 'online_mode',
                'max_players', 'online_players', 'motd_clean', 'country_code',
                'country_name', 'isp', 'first_seen', 'last_seen', 'latency_ms'
            ]
            
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                
                for server in data:
                    # Flatten nested data and ensure all columns exist
                    row = {}
                    for col in columns:
                        row[col] = server.get(col, '')
                    writer.writerow(row)
            
            logger.info(f"Exported {len(data)} servers to CSV: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return False
    
    async def export_excel(self, filename: str, filters: Optional[Dict[str, Any]] = None) -> bool:
        """Export data to Excel format"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl is required for Excel export")
            return False
        
        try:
            data = await self._get_export_data(filters)
            
            if not data:
                logger.warning("No data to export")
                return False
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Servers sheet
            ws_servers = wb.active
            ws_servers.title = "Servers"
            
            # Headers
            headers = [
                'IP', 'Port', 'Version', 'Software', 'Online Mode',
                'Max Players', 'Online Players', 'MOTD', 'Country',
                'ISP', 'First Seen', 'Last Seen', 'Latency (ms)'
            ]
            
            for col, header in enumerate(headers, 1):
                ws_servers.cell(row=1, column=col, value=header)
            
            # Data rows
            for row, server in enumerate(data, 2):
                ws_servers.cell(row=row, column=1, value=server.get('ip', ''))
                ws_servers.cell(row=row, column=2, value=server.get('port', ''))
                ws_servers.cell(row=row, column=3, value=server.get('minecraft_version', ''))
                ws_servers.cell(row=row, column=4, value=server.get('server_software', ''))
                ws_servers.cell(row=row, column=5, value=server.get('online_mode', ''))
                ws_servers.cell(row=row, column=6, value=server.get('max_players', ''))
                ws_servers.cell(row=row, column=7, value=server.get('online_players', ''))
                ws_servers.cell(row=row, column=8, value=server.get('motd_clean', ''))
                ws_servers.cell(row=row, column=9, value=server.get('country_name', ''))
                ws_servers.cell(row=row, column=10, value=server.get('isp', ''))
                ws_servers.cell(row=row, column=11, value=server.get('first_seen', ''))
                ws_servers.cell(row=row, column=12, value=server.get('last_seen', ''))
                ws_servers.cell(row=row, column=13, value=server.get('latency_ms', ''))
            
            # Create summary sheet
            ws_summary = wb.create_sheet("Summary")
            await self._create_summary_sheet(ws_summary, data)
            
            # Save workbook
            wb.save(filename)
            
            logger.info(f"Exported {len(data)} servers to Excel: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return False
    
    async def _get_export_data(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """Get data for export with optional filtering"""
        try:
            # Use the database manager's list_servers or similar method
            # For now, use list_servers and filter in Python if needed
            if hasattr(self.db, 'list_servers'):
                all_servers = await self.db.list_servers()
            else:
                all_servers = []
            if not filters:
                return all_servers
            # Apply filters in Python
            def match(server):
                if 'version' in filters and filters['version'] not in server.get('minecraft_version', ''):
                    return False
                if 'software' in filters and filters['software'] != server.get('server_software', ''):
                    return False
                if 'online_mode' in filters and filters['online_mode'] != server.get('online_mode', ''):
                    return False
                if 'country' in filters and filters['country'] != server.get('country_code', ''):
                    return False
                return True
            return [s for s in all_servers if match(s)]
        except Exception as e:
            logger.error(f"Failed to get export data: {e}")
            return []
    
    async def _create_summary_sheet(self, worksheet, data: List[Dict[str, Any]]) -> None:
        """Create summary statistics sheet"""
        try:
            # Calculate statistics
            total_servers = len(data)
            
            # Count by software
            software_counts = {}
            version_counts = {}
            country_counts = {}
            online_mode_counts = {}
            
            for server in data:
                software = server.get('server_software', 'Unknown')
                software_counts[software] = software_counts.get(software, 0) + 1
                
                version = server.get('minecraft_version', 'Unknown')
                version_counts[version] = version_counts.get(version, 0) + 1
                
                country = server.get('country_name', 'Unknown')
                country_counts[country] = country_counts.get(country, 0) + 1
                
                online_mode = server.get('online_mode', 'Unknown')
                online_mode_counts[online_mode] = online_mode_counts.get(online_mode, 0) + 1
            
            # Write summary data
            row = 1
            
            # Total count
            worksheet.cell(row=row, column=1, value="Total Servers")
            worksheet.cell(row=row, column=2, value=total_servers)
            row += 2
            
            # Software breakdown
            worksheet.cell(row=row, column=1, value="Server Software")
            row += 1
            for software, count in sorted(software_counts.items(), key=lambda x: x[1], reverse=True):
                worksheet.cell(row=row, column=1, value=software)
                worksheet.cell(row=row, column=2, value=count)
                worksheet.cell(row=row, column=3, value=f"{(count/total_servers)*100:.1f}%")
                row += 1
            
            row += 1
            
            # Version breakdown (top 10)
            worksheet.cell(row=row, column=1, value="Top Versions")
            row += 1
            for version, count in sorted(version_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
                worksheet.cell(row=row, column=1, value=version)
                worksheet.cell(row=row, column=2, value=count)
                worksheet.cell(row=row, column=3, value=f"{(count/total_servers)*100:.1f}%")
                row += 1
            
            row += 1
            
            # Online mode breakdown
            worksheet.cell(row=row, column=1, value="Online Mode")
            row += 1
            for mode, count in sorted(online_mode_counts.items(), key=lambda x: x[1], reverse=True):
                worksheet.cell(row=row, column=1, value=mode)
                worksheet.cell(row=row, column=2, value=count)
                worksheet.cell(row=row, column=3, value=f"{(count/total_servers)*100:.1f}%")
                row += 1
            
        except Exception as e:
            logger.error(f"Failed to create summary sheet: {e}")
    
    async def export_players(self, filename: str, filters: Optional[Dict[str, Any]] = None) -> bool:
        """Export player data to CSV or JSON"""
        try:
            if hasattr(self.db, 'list_players'):
                all_players = await self.db.list_players()
            else:
                all_players = []
            if not filters:
                filtered = all_players
            else:
                def match(player):
                    if 'name' in filters and filters['name'] not in player.get('last_known_name', ''):
                        return False
                    return True
                filtered = [p for p in all_players if match(p)]
            # Choose format by file extension
            if filename.endswith('.json'):
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(filtered, f, indent=2, ensure_ascii=False, default=str)
            else:
                columns = ['uuid', 'last_known_name', 'first_seen', 'last_seen', 'total_servers_seen']
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=columns)
                    writer.writeheader()
                    for player in filtered:
                        row = {col: player.get(col, '') for col in columns}
                        writer.writerow(row)
            logger.info(f"Exported {len(filtered)} players to {filename}")
            return True
        except Exception as e:
            logger.error(f"Player export failed: {e}")
            return False

    async def export_mods(self, filename: str, filters: Optional[Dict[str, Any]] = None) -> bool:
        """Export mod data to CSV or JSON"""
        try:
            if hasattr(self.db, 'list_mods'):
                all_mods = await self.db.list_mods()
            else:
                all_mods = []
            if not filters:
                filtered = all_mods
            else:
                def match(mod):
                    if 'id' in filters and filters['id'] not in mod.get('mod_id', ''):
                        return False
                    if 'type' in filters and filters['type'] != mod.get('mod_type', ''):
                        return False
                    return True
                filtered = [m for m in all_mods if match(m)]
            # Choose format by file extension
            if filename.endswith('.json'):
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(filtered, f, indent=2, ensure_ascii=False, default=str)
            else:
                columns = ['mod_id', 'mod_name', 'mod_type', 'first_seen']
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=columns)
                    writer.writeheader()
                    for mod in filtered:
                        row = {col: mod.get(col, '') for col in columns}
                        writer.writerow(row)
            logger.info(f"Exported {len(filtered)} mods to {filename}")
            return True
        except Exception as e:
            logger.error(f"Mod export failed: {e}")
            return False 